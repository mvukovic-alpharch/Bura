from __future__ import annotations

import csv
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
import hashlib
import json
from pathlib import Path
import re
import shutil
import unicodedata


ALIASES = {
    "name": {"name", "player", "player name"}, "dk_player_id": {"id", "player id", "dk player id"},
    "name_id": {"name + id", "name+id"}, "positions": {"position", "roster position", "roster positions"},
    "salary": {"salary"}, "team": {"team", "teamabbrev", "team abbrev"},
    "opponent": {"opponent", "opp"}, "game_info": {"game info", "game"},
    "status": {"status", "injury status"}, "slate_id": {"slate id", "slate_id"},
    "ownership": {"ownership", "projected ownership", "ownership %"},
}
VALID_POSITIONS = {"GK", "G", "D", "M", "F", "UTIL"}


def normalize_name(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode().lower()
    value = re.sub(r"\b(jr|sr|ii|iii|iv)\b", "", value)
    return re.sub(r"[^a-z0-9]", "", value)


def _initial_surname(value: str) -> tuple[str,str] | None:
    plain=unicodedata.normalize("NFKD",value or "").encode("ascii","ignore").decode().lower()
    tokens=re.findall(r"[a-z]+",plain)
    return (tokens[0][0],tokens[-1]) if len(tokens)>=2 else None


def _canonical_header(value: str) -> str:
    key = re.sub(r"[_\-]+", " ", str(value).strip().lower())
    return next((canonical for canonical, aliases in ALIASES.items() if key in aliases), key.replace(" ", "_"))


def _read_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX import requires openpyxl; install project dependencies") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(x or "") for x in next(values)]
        return [dict(zip(headers, row)) for row in values if any(x is not None for x in row)]
    raise ValueError("DraftKings import must be .csv or .xlsx")


def _positions(value) -> list[str]:
    positions = [x.strip().upper() for x in re.split(r"[/,; ]+", str(value or "")) if x.strip()]
    positions = ["GK" if x == "G" else x for x in positions]
    invalid = set(positions) - VALID_POSITIONS
    if not positions or invalid:
        raise ValueError(f"invalid roster eligibility: {value!r}")
    return list(dict.fromkeys(positions))


@dataclass
class MatchResult:
    api_football_player_id: int | None
    internal_player_id: str | None
    confidence: float
    method: str
    review: bool = False


def match_player(row: dict, candidates: list[dict], overrides: dict | None = None) -> MatchResult:
    overrides = overrides or {}
    override = overrides.get(str(row.get("dk_player_id"))) or overrides.get(normalize_name(row["name"]))
    if override is not None:
        player_id = int(override if not isinstance(override, dict) else override["api_football_player_id"])
        return MatchResult(player_id, f"af:{player_id}", 1.0, "manual_override")
    target, team = normalize_name(row["name"]), str(row.get("team") or "").upper()
    exact = [c for c in candidates if normalize_name(c["player_name"]) == target]
    exact_by_id={int(c["player_id"]):c for c in exact}
    if len(exact_by_id) == 1:
        c = next(iter(exact_by_id.values())); pid = int(c["player_id"])
        return MatchResult(pid, f"af:{pid}", 1.0 if not team or team == str(c.get("team") or "").upper() else .97,
                           "exact_name_team" if team and team == str(c.get("team") or "").upper() else "normalized_name")
    if len(exact_by_id) > 1:
        team_exact = {int(c["player_id"]):c for c in exact if team and team == str(c.get("team") or "").upper()}
        if len(team_exact) == 1:
            c=next(iter(team_exact.values())); pid = int(c["player_id"]); return MatchResult(pid, f"af:{pid}", .99, "team_assisted")
        return MatchResult(None, None, 0.0, "ambiguous", True)
    abbreviation=_initial_surname(row["name"])
    if abbreviation:
        abbreviated=[c for c in candidates if _initial_surname(c["player_name"])==abbreviation]
        same_team={int(c["player_id"]):c for c in abbreviated if team and team==str(c.get("team") or "").upper()}
        unique={int(c["player_id"]):c for c in abbreviated}
        pool=same_team if len(same_team)==1 else unique if len(unique)==1 else {}
        if len(pool)==1:
            pid=next(iter(pool)); return MatchResult(pid,f"af:{pid}",.98,"team_initial_surname" if same_team else "initial_surname")
    best_by_id={}
    for c in candidates:
        value=SequenceMatcher(None,target,normalize_name(c["player_name"])).ratio()+(.04 if team and team==str(c.get("team") or "").upper() else 0)
        if int(c["player_id"]) not in best_by_id or value>best_by_id[int(c["player_id"])][0]: best_by_id[int(c["player_id"])]=(value,c)
    scored=sorted(best_by_id.values(),reverse=True,key=lambda x:x[0])
    if not scored:
        return MatchResult(None, None, 0.0, "unmatched", True)
    score, candidate = scored[0]
    second = scored[1][0] if len(scored) > 1 else 0
    if score >= .94 and score-second >= .04:
        pid=int(candidate["player_id"]); return MatchResult(pid, f"af:{pid}", min(score,1.0), "team_assisted_fuzzy")
    return MatchResult(None, None, min(score,1.0), "low_confidence" if score < .94 else "ambiguous", True)


def _candidates(con) -> list[dict]:
    rows = con.execute("""SELECT player_id, any_value(player_name), arg_max(team,date)
        FROM player_match WHERE player_id IS NOT NULL GROUP BY player_id""").fetchall()
    squad_rows=con.execute("SELECT player_id,any_value(player_name),arg_max(team_code,observed_at) FROM player_squads GROUP BY player_id").fetchall()
    combined=[{"player_id":x[0],"player_name":x[1],"team":x[2]} for x in rows if x[0] is not None]
    combined += [{"player_id":x[0],"player_name":x[1],"team":x[2]} for x in squad_rows if x[0] is not None]
    combined += [{"player_id":x[0],"player_name":x[1],"team":None} for x in con.execute("SELECT player_id,player_name FROM players").fetchall() if x[0] is not None]
    unique={(int(x["player_id"]),normalize_name(x["player_name"]),str(x.get("team"))):x for x in combined}
    return list(unique.values())


def ingest_dk_file(con, path: Path, raw_dir: Path, processed_dir: Path, slate_id: str | None = None,
                   salary_cap: int = 50000, overrides_path: Path | None = None) -> dict:
    path = Path(path)
    if not path.is_file(): raise FileNotFoundError(path)
    imported_at = datetime.now(timezone.utc)
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    archive_dir = raw_dir / "draftkings" / imported_at.strftime("%Y/%m/%d")
    archive_dir.mkdir(parents=True, exist_ok=True)
    archived = archive_dir / f"{imported_at.strftime('%Y%m%dT%H%M%S%fZ')}_{digest[:12]}{path.suffix.lower()}"
    shutil.copy2(path, archived)
    raw_rows = _read_rows(archived)
    overrides = json.loads(overrides_path.read_text(encoding="utf-8")) if overrides_path and overrides_path.exists() else {}
    candidates = _candidates(con)
    normalized, errors, duplicates = [], [], []
    seen = set()
    for number, original in enumerate(raw_rows, 2):
        row = {_canonical_header(k): v for k,v in original.items()}
        if not row.get("name") and row.get("name_id"):
            row["name"] = re.sub(r"\s*\(\d+\)\s*$", "", str(row["name_id"])).strip()
        if not row.get("dk_player_id") and row.get("name_id"):
            found = re.search(r"\((\d+)\)\s*$", str(row["name_id"])); row["dk_player_id"] = found.group(1) if found else None
        try:
            salary = int(float(str(row.get("salary") or "").replace("$","").replace(",","")))
            if salary <= 0 or salary > salary_cap: raise ValueError(f"salary {salary} outside 1..{salary_cap}")
            eligibility = _positions(row.get("positions"))
            if not row.get("name"): raise ValueError("missing player name")
        except ValueError as exc:
            errors.append({"row":number,"error":str(exc)}); continue
        dkid = str(row.get("dk_player_id") or f"name:{normalize_name(str(row['name']))}:{row.get('team','')}")
        key = dkid
        if key in seen: duplicates.append({"row":number,"key":key}); continue
        seen.add(key)
        clean={"dk_player_id":dkid,"name":str(row["name"]).strip(),"normalized_name":normalize_name(str(row["name"])),
               "positions":eligibility,"salary":salary,"team":str(row.get("team") or "").upper(),
               "opponent":str(row.get("opponent") or "").upper(),"game_info":str(row.get("game_info") or ""),
               "status":str(row.get("status") or ""),"ownership":float(row["ownership"]) if row.get("ownership") not in (None,"") else None}
        clean["match"] = match_player(clean,candidates,overrides)
        normalized.append(clean)
    if not slate_id:
        supplied = next((str(r.get("slate_id")) for r in ({_canonical_header(k):v for k,v in x.items()} for x in raw_rows) if r.get("slate_id")),None)
        slate_id = supplied or f"dk-{imported_at.strftime('%Y%m%dT%H%M%SZ')}"
    source=str(archived)
    con.execute("BEGIN TRANSACTION")
    try:
        con.execute("DELETE FROM dfs_player_eligibility WHERE slate_id=?",[slate_id])
        con.execute("DELETE FROM dfs_players WHERE slate_id=?",[slate_id])
        con.execute("DELETE FROM dfs_slates WHERE slate_id=?",[slate_id])
        con.execute("INSERT INTO dfs_slates VALUES (?,?,?,?,?,?,?)",[slate_id,imported_at,source,digest,salary_cap,json.dumps(["GK","D","D","M","M","F","F","UTIL"]),"review_required" if errors or duplicates else "imported"])
        for row in normalized:
            m=row["match"]
            con.execute("INSERT INTO dfs_players VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",[slate_id,row["dk_player_id"],m.internal_player_id,m.api_football_player_id,row["name"],row["normalized_name"],row["team"],row["opponent"],row["game_info"],row["salary"],row["status"],row["ownership"],m.confidence,m.method,imported_at,source])
            for position in row["positions"]: con.execute("INSERT INTO dfs_player_eligibility VALUES (?,?,?,?,?)",[slate_id,row["dk_player_id"],position,imported_at,source])
            con.execute("INSERT INTO player_mappings VALUES (?,?,?,?,?,?,?,?,?,?,?)",[m.internal_player_id,m.api_football_player_id,row["dk_player_id"],row["name"],row["normalized_name"],row["team"],"/".join(row["positions"]),m.confidence,m.method,m.method=="manual_override",imported_at])
        con.execute("COMMIT")
    except Exception:
        con.execute("ROLLBACK")
        raise
    matched=sum(x["match"].api_football_player_id is not None for x in normalized)
    ambiguous=[{"dk_player_id":x["dk_player_id"],"name":x["name"],"team":x["team"],"confidence":x["match"].confidence,"reason":x["match"].method} for x in normalized if x["match"].review]
    result={"slate_id":slate_id,"imported_at":imported_at.isoformat(),"source_archive":source,"source_sha256":digest,
            "rows":len(normalized),"matched":matched,"unmatched":len(normalized)-matched,"ambiguous":ambiguous,
            "duplicates":duplicates,"validation_errors":errors}
    report_dir=processed_dir/"reconciliation"; report_dir.mkdir(parents=True,exist_ok=True)
    report=report_dir/f"{slate_id}_{imported_at.strftime('%Y%m%dT%H%M%SZ')}.md"
    report.write_text(_report(result),encoding="utf-8"); result["report"]=str(report)
    return result


def _report(result: dict) -> str:
    lines=[f"# DraftKings reconciliation: {result['slate_id']}","",f"Imported: {result['imported_at']}",f"Matched: {result['matched']} / {result['rows']}",f"Unmatched: {result['unmatched']}",f"Duplicates ignored: {len(result['duplicates'])}",f"Invalid rows ignored: {len(result['validation_errors'])}","","## Manual review"]
    lines += [f"- {x['name']} ({x['team']}): {x['reason']} ({x['confidence']:.3f})" for x in result["ambiguous"]] or ["- None"]
    return "\n".join(lines)+"\n"
